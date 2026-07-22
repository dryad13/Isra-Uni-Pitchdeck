"""M06 — Answer-key file parser (CSV / Excel).

Parses an uploaded answer-key file into normalized `(question_no, option)` pairs.
Tolerant of common column namings and a header-less two-column layout. Question
numbers are interpreted as GLOBAL question numbers (program master-key space).
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

ALLOWED_OPTIONS = {"A", "B", "C", "D", "E"}

_QUESTION_ALIASES = {
    "question", "question_no", "questionno", "q", "qno", "q_no", "no",
    "number", "global_q", "global_question_no", "globalq",
}
_ANSWER_ALIASES = {
    "answer", "correct", "correct_option", "correctoption", "option",
    "key", "ans", "correct_answer",
}


class AnswerKeyParseError(ValueError):
    """Raised when an answer-key file cannot be parsed."""


def _norm(s: str) -> str:
    return s.strip().lower().replace(" ", "_")


def _normalize_option(raw: str) -> str:
    opt = raw.strip().upper()
    # tolerate "1/2/3/4" → A/B/C/D
    digit_map = {"1": "A", "2": "B", "3": "C", "4": "D", "5": "E"}
    if opt in digit_map:
        opt = digit_map[opt]
    if opt not in ALLOWED_OPTIONS:
        raise AnswerKeyParseError(f"Invalid option value: {raw!r} (expected A-E).")
    return opt


def _rows_from_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader if any(cell.strip() for cell in row)]
    except csv.Error as exc:
        raise AnswerKeyParseError(
            "File is not a valid CSV. For a scanned answer sheet, upload JPG/PNG/TIFF "
            "via “Upload marked answer sheet”."
        ) from exc


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise AnswerKeyParseError("openpyxl not installed for Excel parsing.") from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        if any(c.strip() for c in cells):
            rows.append(cells)
    wb.close()
    return rows


_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".pdf", ".gif", ".webp", ".bmp"}


def parse_answer_key(data: bytes, filename: str) -> list[tuple[int, str]]:
    name = filename.lower()
    if Path(filename).suffix.lower() in _IMAGE_SUFFIXES:
        raise AnswerKeyParseError(
            "This looks like a scan/image, not a CSV. Use “Upload marked answer sheet” "
            "or upload a .csv / .xlsx file for spreadsheet keys."
        )
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        rows = _rows_from_xlsx(data)
    else:
        rows = _rows_from_csv(data)

    if not rows:
        raise AnswerKeyParseError("File is empty.")

    header = [_norm(c) for c in rows[0]]
    q_idx = next((i for i, h in enumerate(header) if h in _QUESTION_ALIASES), None)
    a_idx = next((i for i, h in enumerate(header) if h in _ANSWER_ALIASES), None)

    if q_idx is not None and a_idx is not None:
        body = rows[1:]
    else:
        # Header-less: assume column 0 = question, column 1 = option.
        q_idx, a_idx = 0, 1
        body = rows
        if len(rows[0]) < 2:
            raise AnswerKeyParseError(
                "Could not detect question/answer columns. Provide headers "
                "(e.g. 'question_no,correct_option') or a two-column file."
            )

    entries: list[tuple[int, str]] = []
    seen: set[int] = set()
    for line_no, row in enumerate(body, start=2):
        if q_idx >= len(row) or a_idx >= len(row):
            continue
        q_raw, a_raw = row[q_idx].strip(), row[a_idx].strip()
        if not q_raw and not a_raw:
            continue
        try:
            q_no = int(float(q_raw))
        except ValueError as exc:
            raise AnswerKeyParseError(
                f"Row {line_no}: question number {q_raw!r} is not an integer."
            ) from exc
        if q_no < 1:
            raise AnswerKeyParseError(f"Row {line_no}: question number must be >= 1.")
        if q_no in seen:
            raise AnswerKeyParseError(f"Row {line_no}: duplicate question {q_no}.")
        seen.add(q_no)
        entries.append((q_no, _normalize_option(a_raw)))

    if not entries:
        raise AnswerKeyParseError("No answer rows found.")
    return entries
